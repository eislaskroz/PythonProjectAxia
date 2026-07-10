
"""
=========================================================
SERVICIO: services/export_service.py
DESCRIPCIÓN:
Exportación de registros seleccionados desde vistas administrativas.

Soporta:
- XLSX si openpyxl está instalado.
- CSV.
- JSON.
- TXT.
- PDF si reportlab está instalado.
=========================================================
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from tkinter import filedialog, messagebox


def _normalizar_registros(registros):
    if not registros:
        return []
    if isinstance(registros, dict):
        return [registros]
    return list(registros)


def _columnas(registros):
    columnas = []
    for registro in registros:
        for campo in registro.keys():
            if campo not in columnas:
                columnas.append(campo)
    return columnas


def exportar_registros_dialogo(registros, nombre_sugerido="exportacion_axia"):
    """
    Exporta únicamente los registros que el usuario eligió/seleccionó.
    """

    registros = _normalizar_registros(registros)
    if not registros:
        messagebox.showwarning("Exportar", "No hay información seleccionada para exportar.")
        return False

    ruta = filedialog.asksaveasfilename(
        title="Exportar información seleccionada",
        defaultextension=".xlsx",
        initialfile=nombre_sugerido,
        filetypes=[
            ("Excel", "*.xlsx"),
            ("CSV", "*.csv"),
            ("PDF", "*.pdf"),
            ("JSON", "*.json"),
            ("Texto", "*.txt"),
        ],
    )
    if not ruta:
        return False

    ruta = Path(ruta)
    columnas = _columnas(registros)

    try:
        if ruta.suffix.lower() == ".xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
            except Exception:
                # Fallback seguro si openpyxl no está instalado.
                ruta = ruta.with_suffix(".csv")

        if ruta.suffix.lower() == ".xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "AXIA"
            ws.append(columnas)
            for celda in ws[1]:
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill("solid", fgColor="1E3A8A")
                celda.alignment = Alignment(horizontal="center")
            for registro in registros:
                ws.append([registro.get(campo, "") for campo in columnas])
            for idx, campo in enumerate(columnas, start=1):
                ancho = max(len(str(campo)), 14)
                for registro in registros:
                    ancho = max(ancho, min(len(str(registro.get(campo, ""))), 60))
                ws.column_dimensions[get_column_letter(idx)].width = ancho + 2
            wb.save(ruta)
        elif ruta.suffix.lower() == ".pdf":
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet

            doc = SimpleDocTemplate(str(ruta), pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
            estilos = getSampleStyleSheet()
            contenido = [Paragraph("Exportación AXIA", estilos["Title"]), Spacer(1, 12)]
            datos_tabla = [columnas]
            for registro in registros:
                datos_tabla.append([str(registro.get(campo, "") or "")[:120] for campo in columnas])
            tabla = Table(datos_tabla, repeatRows=1)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            contenido.append(tabla)
            doc.build(contenido)
        elif ruta.suffix.lower() == ".json":
            ruta.write_text(json.dumps(registros, ensure_ascii=False, indent=2), encoding="utf-8")
        elif ruta.suffix.lower() == ".txt":
            lineas = []
            for i, registro in enumerate(registros, start=1):
                lineas.append(f"REGISTRO {i}")
                lineas.append("=" * 60)
                for campo in columnas:
                    lineas.append(f"{campo}: {registro.get(campo, '')}")
                lineas.append("")
            ruta.write_text("\n".join(lineas), encoding="utf-8")
        else:
            with ruta.open("w", newline="", encoding="utf-8-sig") as archivo:
                writer = csv.DictWriter(archivo, fieldnames=columnas)
                writer.writeheader()
                for registro in registros:
                    writer.writerow({campo: registro.get(campo, "") for campo in columnas})

        messagebox.showinfo("Exportar", f"Información exportada correctamente:\n{ruta}")
        return True
    except Exception as error:
        messagebox.showerror("Exportar", f"No se pudo exportar la información.\n\n{error}")
        return False
